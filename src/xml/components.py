from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (Alignment, Font, PatternFill)
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText
from openpyxl.chart import Reference


def draw_title(ws: Worksheet,
               crds: tuple,
               title: str | CellRichText,
               font: Font,
               width: int = None,
               height: int = None,
               alignment: Alignment = None,
               sticky: bool = False,
               sticky_offset: int = 1,
               background_color: str = None,
               text_color: str = None) -> tuple:

    cell = ws.cell(row=crds[1], column=crds[0], value=title)
    cell.font = font
    if alignment:
        cell.alignment = alignment
    if sticky:
        letter = get_column_letter(crds[0])
        coord = f'{letter}{crds[1] + sticky_offset}'
        ws.freeze_panes = coord
    if height:
        ws.row_dimensions[crds[1]].height = height
    if background_color:
        cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type='solid')
    if text_color:
        cell.font.color = text_color
    if width:
        ws.merge_cells(start_row=crds[1], start_column=1, end_row=crds[1], end_column=width)
    return crds[0], crds[1] + 1


def draw_row(ws: Worksheet,
             crds: tuple,
             values: list,
             v_offset: int = 0,
             h_offset: int = 0,
             ignore_h_offset: bool = True,
             alignment: Alignment = None):
    for idx, val in enumerate(values, start=crds[0] + h_offset):
        cell = ws.cell(row=crds[1] + v_offset, column=idx, value=val)
        if alignment:
            cell.alignment = alignment

    if not ignore_h_offset:
        return crds[0] + h_offset, crds[1] + 1 + v_offset
    return crds[0], crds[1] + 1 + v_offset


def fill_cell(ws: Worksheet, crds: tuple, color: str, pattern_type: str = 'solid'):
    cell = f'{get_column_letter(crds[0])}{crds[1]}'
    ws[cell].fill = PatternFill(patternType=pattern_type, fgColor=color)
    return cell


def get_refs(ws: Worksheet, crds: tuple, width: int, height: int) -> Reference:
    return Reference(worksheet=ws, min_col=crds[0], max_col=crds[0] + width - 1,
                     min_row=crds[1], max_row=crds[1] + height)
