from typing import List
from openpyxl.chart import Reference, LineChart, ScatterChart
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.worksheet.worksheet import Worksheet
from src.xml.components import draw_title, draw_row, fill_cell, get_refs
from src.xml.styles import H2, H3
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import Marker
from openpyxl.cell.text import Font
from openpyxl.chart.label import DataLabelList


def draw_chart_head(ws: Worksheet, crds: tuple, dates: list, data: dict) -> tuple[tuple, tuple, tuple]:
    heads = dates[:]
    heads.insert(0, '')
    head_crds = draw_row(ws, crds, heads, v_offset=1, alignment=Alignment(horizontal='center', vertical='center'))

    keys = list(data.keys())
    home_row = data[keys[0]]
    away_row = data[keys[1]]

    home_crds = draw_row(ws, head_crds, home_row, h_offset=1, ignore_h_offset=False)
    away_crds = draw_row(ws, home_crds, away_row, ignore_h_offset=True)

    h_cell = fill_cell(ws, (home_crds[0] - 1, home_crds[1] - 1), color='008000')
    a_cell = fill_cell(ws, (away_crds[0] - 1, away_crds[1] - 1), color='B22222')
    ws[h_cell] = keys[0].upper()
    ws[a_cell] = keys[1].upper()
    ws[h_cell].alignment = Alignment(horizontal='center', vertical='center')
    ws[a_cell].alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='Roboto', sz=12, color='ffffff')
    ws[h_cell].font = font
    ws[a_cell].font = font
    return head_crds, home_crds, away_crds


def draw_chart_head_points(ws: Worksheet, crds: tuple, dates: list, data: dict, points: tuple) -> tuple[tuple, tuple, tuple]:
    heads = dates[:]
    heads.insert(0, '')
    head_crds = draw_row(ws, crds, heads, v_offset=1, alignment=Alignment(horizontal='center', vertical='center'))
    ws.column_dimensions['A'].width = 20
    keys = list(data.keys())
    home_row = data[keys[0]]
    away_row = data[keys[1]]

    home_crds = draw_row(ws, head_crds, home_row, h_offset=1, ignore_h_offset=False)
    away_crds = draw_row(ws, home_crds, away_row, ignore_h_offset=True)

    h_cell = fill_cell(ws, (home_crds[0] - 1, home_crds[1] - 1), color='008000')
    a_cell = fill_cell(ws, (away_crds[0] - 1, away_crds[1] - 1), color='B22222')
    ws[h_cell] = f"{keys[0].upper()} {points[0]}"
    ws[a_cell] = f"{keys[1].upper()} {points[1]}"
    ws[h_cell].alignment = Alignment(horizontal='center', vertical='center')
    ws[a_cell].alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='Roboto', sz=12, color='ffffff')
    ws[h_cell].font = font
    ws[a_cell].font = font
    return head_crds, home_crds, away_crds


def draw_chart(data_y: List[Reference],
               crds: tuple,
               data_x: Reference,
               colors: list,
               pos_letter: str,
               size: tuple = (10, 20),
               title: str = "",
               with_labels: bool = True) -> (LineChart, str):
    line_chart = LineChart()
    line_chart.height = size[0]
    line_chart.width = size[1]
    line_chart.title = title
    line_chart.y_axis.title = 'Коэфф.'
    line_chart.x_axis.title = 'Время'
    line_chart.legend = None
    line_chart.x_axis.delete = False
    line_chart.y_axis.delete = False
    line_chart.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(prstDash='dash')))
    line_chart.x_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(prstDash='dash')))

    for idx, dy in enumerate(data_y):
        line_chart.add_data(dy, from_rows=True)
        line = line_chart.series[idx]
        line.marker.size = 6
        line.marker.symbol = 'auto'
        if with_labels:
            line.dLbls = DataLabelList()
            line.dLbls.showCatName = False
            line.dLbls.showLeaderLines = False
            line.dLbls.showLegendKey = False
            line.dLbls.showPercent = False
            line.dLbls.showBubbleSize = False
            line.dLbls.showSerName = False
            line.dLbls.showVal = True
            line.dLbls.position = 'b'
            cp = CharacterProperties(sz=800, b=False)
            line.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp))])
    for idx, color in enumerate(colors):
        line = line_chart.series[idx]
        line.graphicalProperties.ln.solidFill = color

    line_chart.set_categories(data_x)
    row_number = crds[1] + 1
    return line_chart, f'{pos_letter}{row_number}'
