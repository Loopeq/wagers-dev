import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Alignment)
from openpyxl.worksheet.worksheet import Worksheet
from src.utils.common import iso_to_msc
from src.xml.charter.charts import create_line_chart
from src.xml.styles import H2, H1
from components import draw_title
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from utils import reforge

BASE_DIR = Path(__file__).resolve().parent.parent.parent
DATA_DIR = BASE_DIR / 'data'
FILE_DIR = Path(__file__).resolve().parent
WORK_DIR = FILE_DIR / 'sheets'
WORK_DIR.mkdir(exist_ok=True)


def make_head(ws: Worksheet, file_io):
    obj = json.load(file_io)

    sport_name = obj['sport_name']
    league_name = obj["league_name"]
    home_name = obj["home_name"]
    away_name = obj["away_name"]
    start_time = obj['start_time']

    home_font = InlineFont(rFont='Roboto', sz=20, color='008000')
    away_font = InlineFont(rFont='Roboto', sz=20, color='B22222')
    common_font = InlineFont(rFont='Roboto', sz=20)

    head_title = CellRichText([TextBlock(common_font, league_name),
                               TextBlock(common_font, ' | '),
                               TextBlock(home_font, home_name),
                               TextBlock(common_font, ' - '),
                               TextBlock(away_font, away_name),
                               TextBlock(common_font, ' ('),
                               TextBlock(common_font, sport_name),
                               TextBlock(common_font, ')')])
    time_title = f'Начало матча: {iso_to_msc(start_time)}'

    crds = draw_title(ws,
                      crds=(1, 1),
                      title=head_title,
                      width=20,
                      height=30,
                      font=H1,
                      alignment=Alignment(horizontal='center', vertical='center'),
                      sticky=True,
                      sticky_offset=2)

    crds = draw_title(ws,
                      crds=crds,
                      title=time_title,
                      width=20,
                      height=20,
                      font=H2,
                      alignment=Alignment(horizontal='center', vertical='center'))
    return crds


def clear_sheet(wb):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        wb.remove(sheet)


def create():
    wb = Workbook()

    for sport_folder in DATA_DIR.iterdir():
        for match in sport_folder.iterdir():
            clear_sheet(wb)
            try:

                files = [file for file in match.iterdir()]
                head = [file for file in files if 'head.json' in str(file)][0]

                with open(head, 'r') as head_io:
                    head_data = json.load(head_io)
                    start_time = iso_to_msc(head_data['start_time'])
                    start_date = start_time.split()[0]
                    date_path = WORK_DIR / start_date
                    date_path.mkdir(exist_ok=True)

                content = [file for file in files if 'content.json' in str(file)][0]
                with open(content, 'r') as content_io:
                    data, dates = reforge(content_io)
                    for key in data.keys():
                        title = str(key).title()
                        ws = wb.create_sheet(title)
                        with open(head, 'r') as head_io:
                            crds = make_head(ws, head_io)
                        create_line_chart(ws=ws, dates=dates, data=data[key], crds=(crds[0], crds[1]+1), title=title)

                wb.save(date_path / f'{match.name}.xlsx')

            except:
                continue


if __name__ == "__main__":
    create()
