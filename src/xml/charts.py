from typing import List

from openpyxl.chart import Reference, LineChart, ScatterChart
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from components import draw_title, draw_row, fill_cell, get_refs
from openpyxl.worksheet.worksheet import Worksheet
from src.xml.styles import H2, H3
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import Marker


def create_chart_head(ws: Worksheet, crds: tuple, dates: list, data: dict) -> tuple[tuple, tuple, tuple]:

    heads = dates[:]
    heads.insert(0, '')

    head_crds = draw_row(ws, crds, heads, v_offset=1, alignment=Alignment(horizontal='center', vertical='center'))

    home_row = data['home']['values']
    away_row = data['away']['values']

    home_crds = draw_row(ws, head_crds, home_row, h_offset=1, ignore_h_offset=False)
    away_crds = draw_row(ws, home_crds, away_row, ignore_h_offset=True)

    fill_cell(ws, (home_crds[0] - 1, home_crds[1]-1), color='008000')
    fill_cell(ws, (away_crds[0] - 1, away_crds[1]-1), color='B22222')

    return head_crds, home_crds, away_crds


def _create_line_chart(data_y: List[Reference], crds: tuple, data_x: Reference, colors: list, p_offset: int,  chart_offset: int = 1,
                       title: str = "") -> (LineChart, str):
    line_chart = LineChart()
    line_chart.title = title
    line_chart.y_axis.title = 'Коэфф.'
    line_chart.x_axis.title = 'Час'
    line_chart.legend = None
    line_chart.y_axis.scaling.logBase = True
    line_chart.x_axis.delete = False
    line_chart.y_axis.delete = False
    line_chart.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(prstDash='dash')))
    line_chart.x_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(prstDash='dash')))

    for idx, dy in enumerate(data_y):
        line_chart.add_data(dy, from_rows=True)
        line = line_chart.series[idx]
        line.marker.size = 8
        line.marker.symbol = 'auto'
    for idx, color in enumerate(colors):
        line = line_chart.series[idx]
        line.graphicalProperties.ln.solidFill = color
    line_chart.set_categories(data_x)
    col_letter = get_column_letter(crds[0] + (p_offset * 6) + (p_offset * chart_offset) - 1)
    row_number = crds[1] + 1
    return line_chart, f'{col_letter}{row_number}'


def create_line_chart(ws: Worksheet, dates: list[str], data: dict,
                      crds: tuple, title: str = 'MoneyLine - Игра'):

    crds = draw_title(ws, crds, title, width=5, font=H3, height=20,
                      alignment=Alignment(horizontal='left', vertical='center'))

    head_crds, home_crds, away_crds = create_chart_head(ws, crds, dates, data)

    home_values_y = get_refs(ws, crds=(home_crds[0], home_crds[1] - 1), width=len(data['home']['values']), height=0)

    away_values_y = get_refs(ws, crds=(away_crds[0], away_crds[1] - 1), width=len(data['away']['values']), height=0)

    values_x = get_refs(ws, crds=(head_crds[0] + 1, head_crds[1] - 1), width=len(dates), height=0)

    chart, anchor = _create_line_chart(data_y=[home_values_y], data_x=values_x, crds=away_crds,
                                       p_offset=0, chart_offset=2, colors=['008000'])
    ws.add_chart(chart, anchor)
    chart, anchor = _create_line_chart(data_y=[away_values_y], data_x=values_x, crds=away_crds,
                                       p_offset=1, chart_offset=2, colors=['B22222'])
    ws.add_chart(chart, anchor)

    chart, anchor = _create_line_chart(data_y=[home_values_y, away_values_y], data_x=values_x, crds=away_crds,
                                       p_offset=2, chart_offset=2, colors=['008000', 'B22222'])
    ws.add_chart(chart, anchor)



